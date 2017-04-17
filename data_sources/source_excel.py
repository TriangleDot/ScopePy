# -*- coding: utf-8 -*-
"""
Created on Sat Nov 14 15:07:37 2015

@author: john
"""


#==============================================================================
#%% Imports
#==============================================================================

# Standard library
import os


# Third party libraries
import numpy as np
import openpyxl as xl

# My libraries
import data_sources.data_source_base_library as DS



#==============================================================================
#%% Constants
#==============================================================================


# Icons
# ============================
Excel_icons = {"EXCEL_WORKBOOK":"data_source_excel",
               "EXCEL_SHEET":"data_source_excel_sheet",
               "EXCEL_RANGE":"data_source_table"}
              

#==============================================================================
#%% Excel Datasource setup class
#==============================================================================
# This class is a wrapper for functions that convert data sources to a node
# structure
#
# For each source defined, the wrapper class must have the following functions:
# * to_node(*parameters) : returns a DatasourceNode derivative class 
#

class Excel_Datasource(DS.DatasourceCreator):
    """
    Creator class for HDF5 files
    
    """
    
    def __init__(self):
        super(Excel_Datasource, self).__init__()
        
        # Name
        self.name = 'excel'
        
        # It's a file
        self.isFile = True
        
        # File extensions are
        self.fileTypes = ['xlsx']
        
    
    def toNode(self,filename):
        """
        Convert Excel file into a DatasourceNode derivative class
        
        Input
        --------
        filename : str
            path/filename of Excel file
            
        Output
        ---------
        node : csv class instance
        
        """
        
        # Validate filename
        # ------------------------
        if not os.path.exists(filename):
            return
            
        # Open file
        # ------------------
        try:
            print("Making Excel node")
            node = Excel_Workbook(filename)
            
            
            print("Tagging Excel node with filename")
            node.sourceFile = filename
            
            # Tag the base node with the creator name
            node.creator = self.name
            
            return node
            
        except Exception as ex:
            # TODO : log some kind of error here
            print(ex)
            return None
            
            
            
    def file_open(self,filename):
        """
        Required class
        
        Inputs
        --------
        filename : str
            full path/filename
            
        Outputs
        -----------
        source_node : DatasourceNode or derivative
            Excel file structure converted to a node structure.
            
        """
        
        return self.toNode(filename)
        
        
#==============================================================================
#%% Excel Nodes
#==============================================================================
# Node types for Excel files and their components

class Excel_Workbook(DS.DatasourceNode):
    
    def __init__(self, name, parent=None):
        
        # Make name field just the filename
        path, filename = os.path.split(name)
        
        super(Excel_Workbook, self).__init__(filename, parent)
        self.isSource = True
        self._icon_name = Excel_icons[self.typeInfo()]
        
        # Link to file
        self.isFile = True
        self.sourceFile = name
        
        
        self.setName(filename)
        
        self.load()
        
        
    def load(self):
        """
        Load Excel workbook and create child nodes for all the sheets
        
        """
        
        # Load the workbook into memory
        self.workbook = xl.load_workbook(self.sourceFile,data_only=True)
        
        # Create child nodes for all sheets
        # =======================================
        # Get sheet names
        sheets = self.workbook.get_sheet_names()
        
        for sheet in sheets:
            self.addChild(Excel_Sheet(sheet,self.workbook.get_sheet_by_name(sheet)))
        
        
    def close(self):
        """
        Closing function
        
        Closes the Excel file properly
        """
        
        pass
        
        
    def typeInfo(self):
        return "EXCEL_WORKBOOK"
        
        
    def data(self):
        """
        Return Excel file handle
        """
        return self.sourceFile
        
        
    def setData(self,file_path):
        """
        Embed link to source HDF5 file into node structure
        
        Inputs
        ---------
        hdf5_file_handle : h5py File class
        
        """
        self.sourceFile = file_path
        
    
    def path(self):
        """
        Return path of node names back to source node
        
        """
        
        pass   
    
#    @property
#    def sourceFile(self):
#        return self._csv_file
    
    
    
    
class Excel_Sheet(DS.DatasourceTable):
    
    def __init__(self, name,sheet, parent=None):
        super(Excel_Sheet, self).__init__(name, parent)
        self._icon_name = Excel_icons[self.typeInfo()]
        
        # Set this node up as a table
        self.tableWrapper = ExcelSheetTableWrapper
        
        # Store the sheet
        self.setName(name)
        self.sheet = sheet
                
                
        
    def typeInfo(self):
        return "EXCEL_SHEET"
        
        
    def path(self):
        """
        Return path of node names back to source node
        
        """
        return "Excel data"
        
        
    def data(self):
        """
        Return an array-like object from the Excel dataset.
        This acts like a numpy array

        Output
        ---------
        dataset: numpy recarray
        
        """        
        return self.sheet


class Excel_Range(DS.DatasourceTable):
    """
    Not fully implemented yet
    
    """
    
    def __init__(self, name,cell_range, parent=None):
        super(Excel_Range, self).__init__(name, parent)
        self._icon_name = Excel_icons[self.typeInfo()]
        
        # Set this node up as a table
        self.tableWrapper = None
        
        
    def typeInfo(self):
        return "EXCEL_RANGE"
        

#==============================================================================
#%% Table wrapper
#==============================================================================        
        
class ExcelSheetTableWrapper(DS.BaseDatasetTableWrapper):
    """
    Wrapper for Excel sheets when sending to the TableEditor
    
    Makes standard functions for accessing rows and columns, inserting, 
    deleting
    
    """
    
    def __init__(self,sheet):
        """
        Create wrapper with the Excel sheet
        
        Inputs
        ---------
        sheet : openpyxl Worksheet
            Single sheet from an Excel workbook
        
        """
        
        # Initialise base class to get standard properties
        super(ExcelSheetTableWrapper, self).__init__()
        
        # Store the sheet
        self.sheet = sheet
        
        # Read only flag - for completness with other wrappers
        # - prevents insertions and deletions
        self.readOnly = False
        
        # Create column headers here once
        self._columnHeaders = [str(n) for n in range(self.columnCount())]

            
        
    def rowCount(self):
        return self.sheet.max_row
        
        
    def columnCount(self):
        """
        Return number of columns 
        
        Note: openpyxl Worksheet class returns max possible columns at the 
        momement = 1025
        """
        return  self.sheet.max_column
            
        
        
    def cell(self,row,col):
        """
        Return contents of sheet element given row and column
        
        """    

        # Note: openpyxl uses rows and columns starting at 1, not 0         
                
        if row<=self.rowCount() and col<=self.columnCount():
            return self.sheet.cell(row=row+1,column=col+1).value
        
        return
        
        
    def setCell(self,row,col,value):
        """
        Set contents of sheet element given row and column
        
        """
  
        if row<=self.rowCount() and col<=self.columnCount():
            self.sheet.cell(row=row+1,column=col+1).value = value
            
            
    
    def insertRows(self,position,rows=1):
        """
        Insert a set of new rows, set to zero
        
        Inputs
        ---------
        position : int
            row index where new rows will be inserted
            
        rows : int
            Number of rows to be inserted
            
        """
        
        pass
        
        
        
    def removeRows(self,position,rows=1):
        """
        Remove a set of rows
        
        Inputs
        ---------
        position : int
            row index where rows will be removed
            
        rows : int
            Number of rows to be deleted
            
        """
        # TODO: can this be done with a Dataframe?
        
        pass
                
    
        
        
    def columnHeaders(self):
        """
        Return the names of the columns
        
        
        """
        
        # TODO make letter headings in Excel style
        
        return self._columnHeaders
        
        
        
    def columnFormat(self,column):
        """
        Return column format string
        
        Inputs
        --------
        column : int
            column number
            
        Outputs
        --------
        format : str
            format string such as "%.4f'. This should be appropriate to the
            data in the column.
            
        """
        
        # TODO fix this
            
        # Default float format    
        return '%.4f'
        
            
            
            
    def getColumn(self,column_index):
        """
        Return a whole column as one numpy array
        
        """
        pass
    
    
    def getColumnByName(self,column_name):
        """
        Get column using the header name
        
        """
        
        headers = self.columnHeaders()
        
        if column_name not in headers:
            return
            
        pass

#==============================================================================
#%% Source declaration
#==============================================================================
# Put the names of any Datasource classes in the __sources__ variable. When
# ScopePy loads the source classes into memory it will create any class in
# the __sources__ variable
# 

__sources__ = [Excel_Datasource]
